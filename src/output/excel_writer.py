import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from src.utils.logger import Logger

logger = Logger()

# ============================================================
# CLÉS DE MISE À JOUR PAR ONGLET
# ============================================================
SHEET_KEY_COLS = {
    "INIT":     ["Filière", "Type de déclaration", "Type de mesure", "Code", "Déclarant"],
    "CORR":     ["Filière", "Type de déclaration", "Type de mesure", "Code", "Déclarant"],
    "STATS":    ["Filière", "Type de déclaration", "Type de mesure", "Code", "Déclarant"],
    "ANALYSE":  ["Filière", "Type de déclaration", "Type de mesure", "Code", "Déclarant"],
    "DASHBOARD":["Filière", "Type de mesure"],
}

def write_multi_sheet_excel(
    sheets_dict: dict,
    output_path: str,
    overwrite: bool = True
) -> None:
    """
    Écrit plusieurs DataFrames dans un fichier Excel multi-onglets.
    
    - overwrite=True  → recrée le fichier complet
    - overwrite=False → update in place (préserve formats, couleurs, règles métier)
    
    Args:
        sheets_dict (dict): {"nom_onglet": DataFrame, ...}
        output_path (str): Chemin du fichier Excel
        overwrite (bool): Si True, overwrite. Si False, update in place.
    """

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(step="EXCEL_WRITE", message=f"Écriture Excel: {output_path}")

    try:
        # ========== UPDATE IN PLACE (préserve formats) ==========
        if output_path.exists() and not overwrite:
            logger.info(step="EXCEL_WRITE", message="Mise à jour in place du fichier existant")
            wb = load_workbook(output_path)

            def update_sheet(ws_name, df, key_cols):
                # Onglet inexistant → création
                if ws_name not in wb.sheetnames:
                    ws = wb.create_sheet(ws_name)
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    return

                ws = wb[ws_name]

                # Récupérer header existant
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                # Ajouter nouvelles colonnes si besoin
                for col in df.columns:
                    if col not in header:
                        ws.cell(row=1, column=len(header) + 1, value=col)
                        header.append(col)

                key_idx = [header.index(k) for k in key_cols if k in header]

                # Mettre à jour lignes existantes ou en ajouter
                for _, row in df.iterrows():
                    match_row = None
                    for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        if all(excel_row[key_idx[j]].value == row[key_cols[j]]
                               for j in range(len(key_idx))):
                            match_row = excel_row
                            break

                    if match_row:
                        for col in df.columns:
                            if col in header:
                                col_idx = header.index(col)
                                match_row[col_idx].value = row[col]
                                ws.cell(
                                    row=match_row[0].row,
                                    column=col_idx + 1,
                                    value=row[col]
                                )
                    else:
                        new_row = [row[col] if col in df.columns else None for col in header]
                        ws.append(new_row)

            for sheet_name, df in sheets_dict.items():
                if df is not None and not df.empty:
                    key_cols = SHEET_KEY_COLS.get(sheet_name, list(df.columns[:2]))
                    update_sheet(sheet_name, df, key_cols)

            wb.save(output_path)

        # ========== CRÉATION COMPLÈTE ==========
        else:
            logger.info(step="EXCEL_WRITE", message="Création nouveau fichier Excel")
            wb = Workbook()
            wb.remove(wb.active)

            for sheet_name, df in sheets_dict.items():
                if df is not None and not df.empty:
                    ws = wb.create_sheet(sheet_name)
                    _write_df_to_sheet(ws, df)

            wb.save(output_path)

        logger.info(step="EXCEL_WRITE", message=f"✓ Fichier Excel généré: {output_path}")

    except Exception as e:
        logger.error(step="EXCEL_WRITE", message="Erreur lors de la création Excel", exc=e)
        raise


def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    """
    Écrit un DataFrame dans une feuille Excel avec headers et auto-largeur.
    """
    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Données
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-largeur colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)